'''
Created on Mar 18, 2019

@author: dwrigley
'''

from openpyxl import load_workbook
import edcutils
import os
import requests
from requests.auth import HTTPBasicAuth
import urllib3
import json
import platform
import getpass
import time

urllib3.disable_warnings(urllib3.exceptions.SubjectAltNameWarning)

#--------------- settings start here -------------------------
excelFolder="."
excelWorksheet="Dataset Catalog"

catalogServer='http://napslxapp01:9085'
# note:  if using LDAP - prefix uid with LDAP_security_group/
uid='Administrator'
pwd='<prompt>'

# ssl certificate can be empty for http connections
sslCert=""

#catalogServer='https://psrl65dsg018.informatica.com:9553'
#uid='dwrigley'

# if your catalog server uses https - get the certificate as a .cer or .pem and reference here
# default uses the cert from the current folder
#sslCert="psrl65dsg018_base64.pem"

#--------------- settings end here  -------------------------
# declare variable as dict type (global)
custAttrDict={}

def main():
    '''
    reads all .xlsx files in the folder <excelFolder>
    '''

    print("""
*************************************************************************************************************
    Note:  this script is provided for demonstration purposes and is not supported by Informatica
           use to learn how things work and modify as you need to

    Also:  it has been tested to work using python 3, and very minimally python 2.7
           the script may fail with python 2.7 (legacy python) - plesae use v 3 (e.g. e.6 or 3.7) if possible
           see the reference to encode("utf-8") for the occurrence that failed with python 2.7 smoketest

           why - because python 2.7 is end of life at the 1/1/2019. 
                 python 3 is far better - unicode support, fstring & many others
*************************************************************************************************************
    """)

    print("import dataset worksheet from excel test script")
    global custAttrDict
    
    # password prompt - if needed
    global pwd
    if pwd=="<prompt>":
        try: 
            pwd = getpass.getpass(prompt="password for " + uid + ":") 
        except Exception as error: 
            print('ERROR', error) 
        else: 
            print('Password entered:', '*' * len(pwd) ) 

    
    # get a list of all custom attributes - to get the right attribute id's to use
    custAttrDict = listCustomAttributes()
    print("")
     
    print("step 2: reading file list from: folder=" + excelFolder)
    excelCount=0
    for xlsFile in os.listdir(excelFolder):
        if xlsFile.endswith('.xlsx'):
            excelCount += 1
            processExcelFile(xlsFile)
    
    print("files found/processed=" + str(excelCount))
    print("finished")


def processExcelFile(excelFile):
    """
    read the excel file - look for the worksheet named 'Dataset Catalog'
    then based on the contents- find and potentially update EDC with custom attribute values

    after reading all data - call a method that will find the object in the catalog & update if necessary
    """
    fqn = excelFolder + "/" + excelFile
    print("\tprocessing file: " + fqn)
    
    wb = load_workbook(filename=fqn, read_only=True)
    print("\t\texcel file opened...")
    print("\t\tworksheet names: "+ str(wb.sheetnames))
    if (excelWorksheet not in wb.sheetnames):
        print("\t\tworksheet not found: " + excelWorksheet + " exiting")
        return
    
    # assume the worksheet is found
    active_ws=wb[excelWorksheet]
    print("\t\topened worksheet...." + excelWorksheet)
    
    # for each row (after the header) - first get the "DataSet Name"
    datasetName=""
    
    """
    get all of the attibute names/values from columns 1 & 3 in the excel workbook
    """
    row=0
    # attrs - this dict will store key=attr name, val=attr val
    attrs={}
    for rows in active_ws.iter_rows(min_row=2, max_row=0, min_col=0):
        row += 1
        cellNum=0
        print("\t\trow: %s" % row)
        for cell in rows:
            cellNum += 1
            if cellNum==1:
                # attribute name
                if (cell.value != None):
                    datasetName=cell.value
                    if str(platform.python_version()).startswith("2.7"):
                        print("\t\tattributeName=" + datasetName.encode("utf-8"))
                    else:
                        print("\t\tattributeName=" + datasetName)
                else:
                    break
            if cellNum==3:
                # Catalog Attribute Value
                if(cell.value != None):
                    attrValue = cell.value
                    print("\t\tvalue=" + str(attrValue));
                    # the example has trailing spaces on some attribute names - strip them
                    attrs[datasetName.strip()]=attrValue
    # end of row
    #print(attrs)
    print("\tfinished reading data from excel - calling findAndUpdateCatalog passing attribute list")
    print("")
    
    # find the object in the catalog and update
    findAndUpdateCatalog(attrs)
    
    
def findAndUpdateCatalog(attrDict):
    '''
    given a list of properties - find an object in the catalog - and compare the attributes found, update if necessary
    '''
    objectName = attrDict["Dataset Name"]
    #objectName = 'CRM_CUSTOMER_MAIN'
    techName = objectName.replace(" ", "_")
    print("\tready to find and update catalog object: " + objectName + " attrs=" + str(attrDict))
    print("\ttechName=" + techName)
    
        
    header = {"Accept": "application/json"} 
    query = "+core.allclassTypes:com.infa.ldm.relational.Table +core.name:" + techName
    parameters = {'offset': "0", 'pageSize': "200", 'q': query}

    # execute catalog rest call, for a page of results
    print('\texecuting search: q=' + query)
    resp = requests.get(catalogServer + '/access/2/catalog/data/objects', params=parameters, headers=header, auth=HTTPBasicAuth(uid,pwd), verify=sslCert)
    status = resp.status_code
    if status != 200:
        # some error - e.g. catalog not running, or bad credentials
        print("error! " + str(status) + str(resp.json()))
        return
        
    resultJson = resp.json()
    total=resultJson['metadata']['totalCount']
    #print(total)
    if total != 1:
        print("\t" + str(total) + " item(s) found - looking for exactly 1 - catalog object will not be updated")
        return

    itemId=""
    itemName=""
    itemHref=""
    # get the eTag
    eTag = resp.headers['ETag']
    print("\titem found: eTag=" + eTag)
    for foundItem in resultJson["items"]:
        itemId=foundItem["id"]
        itemName=edcutils.getFactValue(foundItem, "core.name")
        itemHref=foundItem["href"]
        
        # get the facts collection (to be updated for any new/upated attributes
        itemFacts = foundItem['facts']
            
        print("\tid=" + itemId)
        print("\tname=" + itemName)
#         print("\tfacts=" + str(itemFacts))
        
        # flag to determine if an updated is necessary
        updateObject=False
        
        # activity counts
        newFactCount=0
        updFactCount=0
        
        for excelAttrName, excelAttrVal in attrDict.items():
            print("\t\tchecking excel attr=" + excelAttrName)
            doesAttrExist = excelAttrName in custAttrDict
#             print("\t\t\tattr exists in EDC? " + str(doesAttrExist))
            if doesAttrExist:
                # check if the content matches
                edcAttrVal=edcutils.getFactValue(foundItem, custAttrDict.get(excelAttrName))
                if (edcAttrVal == None):
                    newFactCount += 1
                    print("\t\t\tnew attribute value...")
                    newFact = {'attributeId': custAttrDict.get(excelAttrName), 'value':  excelAttrVal, 'label': excelAttrName}
                    print("\t\t\t\t\t" + str(newFact))
                    itemFacts.append(newFact)
                    updateObject=True
                                        
                else:
                    print("\t\t\texisting value='" + edcAttrVal + "'=='" + excelAttrVal + "' " + str(edcAttrVal == excelAttrVal) )
                    # 
                    # compare the values - if match, do nothing
                    # if different - update
                    if (edcAttrVal == excelAttrVal):
                        print("\t\t\tvalues match - no update needed")
                    else:
                        updFactCount += 1
                        updateObject = True
                        # update the actual value...
                        print("\t\t\treplacing: attr=" + custAttrDict.get(excelAttrName) + " from=" + edcAttrVal + " with " + excelAttrVal)
                        replaceExistingAttrValue(foundItem, custAttrDict.get(excelAttrName), excelAttrVal)
            else:
                # attribute does not exist
                print("\t\t\tattr '" +excelAttrName + "' does not exist - use ldmadmin ui to create it")           

                
        # print message if nothing to do & return  
        if not updateObject:
            print("\tno updates needed for object")
            return
        
        # object should be updated
        print("\tobject should be updated - new facts:" + str(newFactCount) + " updated facts:" + str(updFactCount))
        
        # uncomment these 2 lines to help test concurrent updates - while the sleep is happening, manually update the object
        # this should cause a 412 error message (ifMatch <> eTag)
        #print("sleeping for 20 seconds - to simulate concurrent updates")
        #time.sleep(20)

        # prepare to write the item to the catalog 
        putParms = {'Content-Type': 'application/json', 'Accept': 'application/json', 'If-Match': eTag}
        putUrl = catalogServer + '/access' + itemHref
        print("\tready to update item: url=" + putUrl + " parm=" + str(putParms))
        # make the PUT request (execute the updaate
        r = requests.put(putUrl, data=json.dumps(foundItem), headers=putParms, auth=HTTPBasicAuth(uid,pwd), verify=sslCert)
        putStatus=r.status_code
        print("\treturn code=" + str(putStatus))
        if putStatus==200:
            newEtag=r.headers['eTag']
            print("\t\trequest completed: status=%d new eTag %s" % (putStatus, newEtag))
        else:
            print("update failed." + str(r.json()))

        


    
    
def listCustomAttributes():
    '''
    query the catalog to get a list of all custom attributes.  
    store in menory (dict:  key=attr name  val=attr id
    
    use this later to verify that the attribute referenced in excel actually exists
    '''
    tmpDict={}
    resturl = catalogServer + '/access/2/catalog/models/attributes'
    print("Step 1: listing custom attributes: " + resturl)
    header = {"Accept": "application/json"} 

    total=1000  # initial value - set to > 0 - will be over-written by the count of objects returned
    offset=0
    page=0
    pageSize=500

    while offset<total:
        page += 1
        parameters = {'offset': offset, 'pageSize': pageSize}

        # execute catalog rest call, for a page of results
        resp = requests.get(resturl, params=parameters, headers=header, auth=HTTPBasicAuth(uid,pwd), verify=sslCert)
        status = resp.status_code
        if status != 200:
            # some error - e.g. catalog not running, or bad credentials
            print("error! " + str(status) + str(resp.json()))
            break
        
        resultJson = resp.json()
        total=resultJson['metadata']['totalCount']
        #print("objects found: " + str(total) + " offset: " + str(offset) + " pagesize="+str(pageSize) + " currentPage=" + str(page) );

        # for next iteration
        offset += pageSize;
        # for each attribute found...
        for attrDef in resultJson["items"]:
            attrId = attrDef["id"]
            attrName = attrDef["name"]
            #dataType = attrDef["dataTypeId"]
            #sortable = attrDef["sortable"]
            #facetable = attrDef["facetable"]
            if attrId.startswith("com.infa.appmodels.ldm."):
                # add to the collection
                tmpDict[attrName]=attrId
                        
    # end of while loop
    print("\tlistCustomAtrs returning : " + str(len(tmpDict)) + " " + str(tmpDict.keys()))
    
    # note:  we shold be returning the dict
    return tmpDict

def replaceExistingAttrValue(theItem, theAttribute, theValue):
    '''
    facts are represented as a list of dictionaries
    each fact will have:  attributeId and value
    for the given attributeId (in 'theAttribute') we want to replace the value with theValue
    '''
    for eachfact in theItem['facts']:
        if eachfact.get('attributeId') == theAttribute:
            #print("before value")
            #print (str(eachfact))
            # update the value - the beauty of python is that this is referenced by the original item json (so it ready to write now)
            eachfact['value']=theValue
            #print("after value")
            #print (str(eachfact))

        



if __name__ == '__main__':
    main()

